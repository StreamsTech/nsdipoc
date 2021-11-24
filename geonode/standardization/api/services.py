import os
import re
import tempfile
from datetime import datetime

from openpyxl import load_workbook

from geonode.data_upload.models import MasterData, BillingData, StatusUpdateModel, DeepTubewellData
from geonode.data_upload.utility_dict import MODEL_DATA_FIELDS_DICT as data_dict, MASTER_DATA_CELL_TYPES, \
    BILLING_DATA_CELL_TYPES, DTW_DATA_CELL_TYPES
from geonode.data_upload.enums import DataTypeEnum
from geonode.data_upload.store_procedure import update_geodata_from_billing_data, update_geodata_from_master_data, update_geodata_from_dtw_data
from geonode.data_upload.store_procedure import calculate_water_history_dtw, calculate_water_history_consumption


def get_data_upload_service(data_type):
    if data_type == DataTypeEnum.MASTER.value[0]:
        return MasterDataUploadService
    if data_type == DataTypeEnum.BILLING.value[0]:
        return BillingDataUploadService
    if data_type == DataTypeEnum.DTW.value[0]:
        return DTWDataUploadService


class BaseDataUploadService:
    def __init__(self, user, zone, dma, data_file):
        self.user = user
        self.zone = zone
        self.dma = dma
        self.data_file = data_file
        self.workbook = load_workbook(filename=data_file, read_only=True, data_only=True)
        self.sheet = self.workbook.active
        self.data_validation_status = []

    @classmethod
    def get_file_path(cls, request):
        file_name = request.FILES['file'].name
        _, file_extension = os.path.splitext(file_name)
        if file_extension != '.xlsx':
            return None, "File extension should be '.xlsx', " + file_extension + " is given"
        temp_dir = tempfile.mkdtemp()
        temporary_file = temp_dir + '/' + file_name
        with open(temporary_file, 'wb+') as f:
            for chunk in request.FILES['file'].chunks():
                f.write(chunk)
        return f.name, None

    @staticmethod
    def clear_existing_data(model, **kwargs):
        model.objects.filter(**kwargs).delete()

    @staticmethod
    def process_row(row):
        values = []
        value_types = []
        for cell in row:
            values.append(cell.value)
            value_types.append(type(cell.value))
        return values


    @staticmethod
    def update_upload_status(error_list, uploaded_file_name, uploader, data_type):
        upload_status = StatusUpdateModel()
        upload_status.file_name = uploaded_file_name
        upload_status.uploader = uploader
        upload_status.file_type = data_type
        if len(error_list) > 0:
            upload_status.upload_status = "Failed"
            upload_status.has_errors = True
        else:
            upload_status.upload_status = "Success"
            upload_status.has_errors = False
        upload_status.errors = error_list  # json.dumps(error_list)
        upload_status.save()

    def validate_row_data(self, row, data_field_types):
        fields_len = len(data_field_types)
        for i in xrange(fields_len):
            cell_type = type(row[i].value)
            db_field_type = data_field_types[i]['type']

            if cell_type == db_field_type:
                continue
            if (cell_type == type(None) or cell_type == '') and data_field_types[i]['is_null']:
                continue
            if cell_type == unicode and db_field_type == datetime:
                continue

            error = {
                'cell': row[i].coordinate,
                'expected': str(data_field_types[i]['type']),
                'given': str(cell_type)
            }
            self.data_validation_status.append(error)


class MasterDataUploadService(BaseDataUploadService):
    def __init__(self, user, zone, dma, data_file):
        self.model_fields = data_dict['MASTER']['fields']
        BaseDataUploadService.__init__(self, user, zone, dma, data_file)

    def is_valid(self):
        if self.sheet.max_row <= 1:
            return False, "The data file you are trying to upload is empty"
        min_num_column = len(self.sheet[1])
        if min_num_column < len(MASTER_DATA_CELL_TYPES):
            return False, "Missing columns for master data"
        return True, ''

    @staticmethod
    def process_customer_category(model_object):
        domestic_value = None
        commercial_value = None
        community_value = None
        values = model_object.customer_category.split(" ")

        if 'Domestic' in values:
            domestic_value = re.sub("[^0-9]", "", values[values.index('Domestic') + 1])
        if 'Commercial' in values:
            commercial_value = re.sub("[^0-9]", "", values[values.index('Commercial') + 1])
        if 'Community' in values:
            community_value = re.sub("[^0-9]", "", values[values.index('Community') + 1])

        model_object.commercial_customer = commercial_value
        model_object.domestic_customer = domestic_value
        model_object.community_customer = community_value

        return model_object

    @staticmethod
    def process_date_fields(model_object):
        if model_object.meter_installation_date and model_object.meter_installation_date != '':
            model_object.meter_installation_date = datetime.strptime(model_object.meter_installation_date, '%d-%b-%Y')
        if model_object.cp_date and model_object.cp_date != '':
            model_object.cp_date = datetime.strptime(model_object.cp_date, '%d-%b-%Y')
        return model_object

    def save_data(self):
        data_list = []
        self.clear_existing_data(MasterData, zone_id=self.zone, dma_id=self.dma)
        for row in self.sheet.iter_rows(min_row=2):
            values = self.process_row(row)
            if not values[0]:
                continue
            data_dict = dict(zip(self.model_fields, values))
            data_dict["user_id"] = self.user.id
            data_dict["zone_id"] = self.zone
            data_dict["dma_id"] = self.dma

            model_object = MasterData(**data_dict)
            model_object = self.process_customer_category(model_object)
            model_object = self.process_date_fields(model_object)
            self.validate_row_data(row, MASTER_DATA_CELL_TYPES)
            data_list.append(model_object)

        if len(self.data_validation_status) == 0:
            MasterData.objects.bulk_create(data_list)
            update_geodata_from_master_data()
        uploaded_file_name = self.data_file.split('/')[-1]
        self.update_upload_status(self.data_validation_status, uploaded_file_name, self.user, 'Master Data')


class BillingDataUploadService(BaseDataUploadService):
    def __init__(self, user, zone, dma, data_file):
        self.model_fields = data_dict['BILLING']['fields']
        BaseDataUploadService.__init__(self, user, zone, dma, data_file)

    def is_valid(self):
        if self.sheet.max_row <= 1:
            return False, "The data file you are trying to upload is empty"
        min_num_column = len(self.sheet[1])
        if min_num_column < len(BILLING_DATA_CELL_TYPES):
            return False, "Missing columns for billing data"
        return True, ''

    @staticmethod
    def process_date_fields(model_object):
        if model_object.previous_date and model_object.previous_date != '':
            model_object.previous_date = datetime.strptime(model_object.previous_date, '%d-%b-%Y')
        if model_object.present_date and model_object.present_date != '':
            model_object.present_date = datetime.strptime(model_object.present_date, '%d-%b-%Y')
        if model_object.issue_date and model_object.issue_date != '':
            model_object.issue_date = datetime.strptime(model_object.issue_date, '%d-%b-%Y')
        if model_object.pay_date and model_object.pay_date != '':
            model_object.pay_date = datetime.strptime(model_object.pay_date, '%d-%b-%Y')
        return model_object

    def save_data(self):
        data_list = []

        self.clear_existing_data(BillingData, year_month=self.sheet[2][5].value, zone_id=self.zone, dma_id=self.dma)
        for row in self.sheet.iter_rows(min_row=2):
            values = self.process_row(row)
            if not values[0]:
                continue
            data_dict = dict(zip(self.model_fields, values))
            data_dict["user_id"] = self.user.id
            data_dict["zone_id"] = self.zone
            data_dict["dma_id"] = self.dma

            model_object = BillingData(**data_dict)
            model_object = self.process_date_fields(model_object)
            self.validate_row_data(row, BILLING_DATA_CELL_TYPES)
            data_list.append(model_object)

        if len(self.data_validation_status) == 0:
            res = BillingData.objects.bulk_create(data_list)
            update_geodata_from_billing_data()
            if res:
                consumption_year_month = res[0].year_month
                calculate_water_history_consumption(consumption_year_month)
        uploaded_file_name = self.data_file.split('/')[-1]
        self.update_upload_status(self.data_validation_status, uploaded_file_name, self.user, 'Billing Data')


class DTWDataUploadService(BaseDataUploadService):
    def __init__(self, user, zone, dma, data_file):
        self.model_fields = data_dict['DTW']['fields']
        BaseDataUploadService.__init__(self, user, zone, dma, data_file)

    def is_valid(self):
        if self.sheet.max_row <= 1:
            return False, "The data file you are trying to upload is empty"
        min_num_column = len(self.sheet[1])
        if min_num_column < len(DTW_DATA_CELL_TYPES):
            return False, "Missing columns for deep tubewell data"
        return True, ''

    def save_data(self):
        data_list = []

        self.clear_existing_data(DeepTubewellData, year_month=self.sheet[2][4].value, zone_id=self.zone, dma_id=self.dma)
        for row in self.sheet.iter_rows(min_row=2):
            values = self.process_row(row)
            if not values[0]:
                continue
            data_dict = dict(zip(self.model_fields, values))
            data_dict["user_id"] = self.user.id
            data_dict["zone_id"] = self.zone
            data_dict["dma_id"] = self.dma

            model_object = DeepTubewellData(**data_dict)
            self.validate_row_data(row, DTW_DATA_CELL_TYPES)
            data_list.append(model_object)

        if len(self.data_validation_status) == 0:
            DeepTubewellData.objects.bulk_create(data_list)
            for dtw_data in data_list:
                year_month = dtw_data.year_month
                dma_id = dtw_data.dma_id
                update_geodata_from_dtw_data(dma_id, year_month)
                calculate_water_history_dtw(dma_id, year_month)
        uploaded_file_name = self.data_file.split('/')[-1]
        self.update_upload_status(self.data_validation_status, uploaded_file_name, self.user, 'DTW Data')
